import csv
import io
import math
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from django.db import connection, transaction
from django.utils import timezone
from openpyxl import load_workbook
from psycopg import sql

from .models import ImportColumnMapping, ImportMappingTemplate, ImportMappingTemplateColumn, RawImportBatch, RawImportRow


IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
TYPE_MAP = {
    "text": "text",
    "bigint": "bigint",
    "integer": "integer",
    "double precision": "double precision",
    "boolean": "boolean",
    "date": "date",
}
PRODUCTION_DAILY_TABLE = "production_daily"
PRODUCTION_MONTHLY_TABLE = "production_monthly"
INJECTION_DAILY_TABLE = "injection_daily"
INJECTION_MONTHLY_TABLE = "injection_monthly"
PRODUCTION_COLUMN_ALIASES = {
    "base_uwi": (
        "base_uwi", "well_id", "well id", "wellid", "uwi", "api", "well", "well_number",
        "well number", "lease_well_id", "lease well id",
    ),
    "production_date": ("date", "production_date", "production date", "prod_date", "prod date", "day"),
    "daily_oil": ("daily_oil", "daily oil", "oil", "oil_volume", "oil volume", "oil_bbl", "oil bbl"),
    "daily_water": ("daily_water", "daily water", "water", "water_volume", "water volume", "water_bbl", "water bbl"),
    "daily_gas": ("daily_gas", "daily gas", "gas", "gas_volume", "gas volume", "gas_mcf", "gas mcf"),
    "fluid": ("fluid", "daily_fluid", "daily fluid", "total_fluid", "total fluid", "liquid", "liquids"),
    "tbg_psi": ("tbg_psi", "tbg psi", "tbg", "tubing pressure", "tubing_pressure", "TP", "tp"),
    "csg_psi": ("csg_psi", "csg psi", "csg", "casing pressure", "casing_pressure", "CP", "cp"),
    "GOR": ("GOR", "gas oil ratio", "gas_oil_ratio"),
    "GLR": ("GLR", "gas liquid ratio", "gas_liquid_ratio"),
}
WELL_HEADER_MOVED_COLUMNS = (
    "cur_operator_code", "cur_operator_name", "lahee_class_code", "lic_subs_well_obj",
    "lic_wa_wid_permit", "org_operator_name", "orig_operator_code", "orig_units_m_api",
    "well_pad_id", "well_type", "Permit_id",
)


def normalize_header(value, position, used):
    original = str(value).strip() if value not in (None, "") else f"column_{position}"
    base = re.sub(r"[^a-z0-9]+", "_", original.casefold()).strip("_")
    if not base:
        base = f"column_{position}"
    if base[0].isdigit():
        base = f"column_{base}"

    candidate = base
    counter = 2
    while candidate in used:
        candidate = f"{base}__{counter}"
        counter += 1
    used.add(candidate)
    return candidate


def header_key(value):
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").casefold()).strip("_")


def json_value(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def load_mapping_template(headers):
    path = Path(settings.BASE_DIR) / "megadata" / "table_column_mapping_working.csv"
    header_lookup = {header.casefold(): header for header in headers}
    mappings = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for order, row in enumerate(csv.DictReader(handle), start=1):
            target_column = (row.get("target_column") or "").strip()
            configured_source = (row.get("source_column") or "").strip()
            source_aliases = {
                "cur_operator_name": ["current_operator", "operator", "last_permit_operator"],
                "cur_operator_code": ["current_operator_code", "operator_number", "last_permit_operator_number"],
                "md_all_wells_m": ["measured_depth_m", "total_depth_m", "total_depth"],
            }
            suggested = header_lookup.get(configured_source.casefold(), "")
            if not suggested and target_column in source_aliases:
                suggested = next(
                    (header_lookup[alias] for alias in source_aliases[target_column] if alias in header_lookup),
                    "",
                )
            mappings.append(
                ImportColumnMapping(
                    mapping_order=order,
                    target_table=(row.get("target_table") or "").strip(),
                    target_column=target_column,
                    target_type=(row.get("target_type") or "text").strip().lower(),
                    suggested_source_column=suggested,
                    source_column=suggested,
                    transform_rule=(row.get("transform_rule") or "").strip(),
                    casing_type=(row.get("casing_type") or "").strip(),
                    include=(row.get("include_flag") or "").strip().upper() == "Y",
                    required=(row.get("required_flag") or "").strip().upper() == "Y",
                    default_value=(row.get("default_value") or "").strip(),
                    key_role=(row.get("key_role") or "").strip(),
                    notes=(row.get("notes") or "").strip(),
                )
            )
    return mappings


@transaction.atomic
def ingest_workbook(uploaded_file, requested_sheet=""):
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("Upload an .xlsx or .xlsm workbook.")

    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise ValueError(f"Worksheet '{requested_sheet}' was not found.")
        worksheet = workbook[requested_sheet]
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    rows = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as exc:
        raise ValueError("The selected worksheet is empty.") from exc

    generated_headers = ["import_timestamp", "raw_id", "source_file"]
    used = set(generated_headers)
    source_headers = [normalize_header(value, index, used) for index, value in enumerate(raw_headers, start=1)]
    headers = [*generated_headers, *source_headers]
    batch = RawImportBatch.objects.create(file_name=uploaded_file.name, sheet_name=worksheet.title, headers=headers)
    import_timestamp = batch.imported_at.isoformat()

    pending = []
    row_count = 0
    for excel_row_number, values in enumerate(rows, start=2):
        values = list(values)
        if not any(value not in (None, "") for value in values):
            continue
        if len(values) < len(source_headers):
            values.extend([None] * (len(source_headers) - len(values)))
        row_count += 1
        generated_raw_id = batch.id * 1_000_000_000 + row_count
        data = {
            "import_timestamp": import_timestamp,
            "raw_id": generated_raw_id,
            "source_file": uploaded_file.name,
            **{header: json_value(values[index]) for index, header in enumerate(source_headers)},
        }
        pending.append(RawImportRow(batch=batch, row_number=excel_row_number, data=data))
        if len(pending) >= 1000:
            RawImportRow.objects.bulk_create(pending, batch_size=1000)
            pending = []
    if pending:
        RawImportRow.objects.bulk_create(pending, batch_size=1000)

    mappings = load_mapping_template(headers)
    for mapping in mappings:
        mapping.batch = batch
    ImportColumnMapping.objects.bulk_create(mappings, batch_size=1000)
    batch.row_count = row_count
    batch.save(update_fields=["row_count"])
    workbook.close()
    return batch


def raw_text_value(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


@transaction.atomic
def materialize_raw_table(batch):
    table_name = f"raw_excel_batch_{batch.id}"
    validate_identifier(table_name)

    metadata_column = "_excel_row_number"
    while metadata_column in batch.headers:
        metadata_column = f"_{metadata_column}"

    with connection.cursor() as cursor:
        cursor.execute(sql.SQL("DROP TABLE IF EXISTS {}").format(sql.Identifier(table_name)))
        raw_column_types = {
            "import_timestamp": "timestamptz",
            "raw_id": "bigint",
            "source_file": "text",
        }
        column_definitions = [
            sql.SQL("{} bigint NOT NULL").format(sql.Identifier(metadata_column)),
            *[
                sql.SQL("{} {}").format(
                    sql.Identifier(validate_identifier(header)),
                    sql.SQL(raw_column_types.get(header, "text")),
                )
                for header in batch.headers
            ],
        ]
        cursor.execute(
            sql.SQL("CREATE TABLE {} ({})").format(
                sql.Identifier(table_name),
                sql.SQL(", ").join(column_definitions),
            )
        )

        columns = [metadata_column, *batch.headers]
        statement = sql.SQL("INSERT INTO {} ({}) VALUES ({})").format(
            sql.Identifier(table_name),
            sql.SQL(", ").join(map(sql.Identifier, columns)),
            sql.SQL(", ").join(sql.Placeholder() for _ in columns),
        )
        pending = []
        for raw_row in batch.rows.order_by("row_number").iterator(chunk_size=1000):
            pending.append([
                raw_row.row_number,
                *[raw_text_value(raw_row.data.get(header)) for header in batch.headers],
            ])
            if len(pending) >= 1000:
                cursor.executemany(statement, pending)
                pending = []
        if pending:
            cursor.executemany(statement, pending)

    batch.raw_table_name = table_name
    batch.raw_table_created_at = timezone.now()
    batch.error_message = ""
    batch.save(update_fields=["raw_table_name", "raw_table_created_at", "error_message"])
    return table_name



def table_column_names(cursor, table_name):
    validate_identifier(table_name)
    cursor.execute(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = %s
        ORDER BY ordinal_position
        """,
        [table_name],
    )
    return [row[0] for row in cursor.fetchall()]


@transaction.atomic
def create_unique_api_table(batch, api_column="api"):
    if not batch.raw_table_name:
        raise ValueError("Create the raw database table before processing unique API values.")

    raw_table_name = validate_identifier(batch.raw_table_name)
    unique_table_name = validate_identifier(f"unique_{raw_table_name}")
    api_column = validate_identifier(api_column)

    with connection.cursor() as cursor:
        columns = table_column_names(cursor, raw_table_name)
        if api_column not in columns:
            raise ValueError(f"The raw table does not contain an '{api_column}' column.")

        select_fields = [
            sql.SQL("NULLIF(BTRIM({}), '') AS {}").format(sql.Identifier(api_column), sql.Identifier(api_column)),
            sql.SQL("COUNT(*) AS source_row_count"),
        ]
        for column_name in columns:
            if column_name == api_column:
                continue
            column = sql.Identifier(column_name)
            normalized = sql.SQL("NULLIF(BTRIM({}::text), '')").format(column)
            select_fields.append(
                sql.SQL("STRING_AGG(DISTINCT {value}, ' | ' ORDER BY {value}) FILTER (WHERE {value} IS NOT NULL) AS {alias}").format(
                    value=normalized,
                    alias=column,
                )
            )

        normalized_api = sql.SQL("NULLIF(BTRIM({}), '')").format(sql.Identifier(api_column))
        cursor.execute(sql.SQL("DROP TABLE IF EXISTS {}").format(sql.Identifier(unique_table_name)))
        cursor.execute(
            sql.SQL(
                "CREATE TABLE {} AS SELECT {} FROM {} WHERE {} IS NOT NULL GROUP BY {} ORDER BY {}"
            ).format(
                sql.Identifier(unique_table_name),
                sql.SQL(", ").join(select_fields),
                sql.Identifier(raw_table_name),
                normalized_api,
                normalized_api,
                normalized_api,
            )
        )
        cursor.execute(sql.SQL("CREATE INDEX {} ON {} ({})").format(
            sql.Identifier(f"{unique_table_name}_api_idx"),
            sql.Identifier(unique_table_name),
            sql.Identifier(api_column),
        ))
        cursor.execute(sql.SQL("SELECT COUNT(*) FROM {}").format(sql.Identifier(unique_table_name)))
        unique_row_count = cursor.fetchone()[0]

    batch.unique_table_name = unique_table_name
    batch.unique_table_created_at = timezone.now()
    batch.unique_row_count = unique_row_count
    batch.error_message = ""
    batch.save(update_fields=["unique_table_name", "unique_table_created_at", "unique_row_count", "error_message"])
    return {
        "table_name": unique_table_name,
        "row_count": unique_row_count,
    }
def serialize_batch(batch, include_mappings=False, preview_limit=20):
    payload = {
        "id": batch.id,
        "file_name": batch.file_name,
        "sheet_name": batch.sheet_name,
        "status": batch.status,
        "headers": batch.headers,
        "row_count": batch.row_count,
        "raw_table_name": batch.raw_table_name,
        "raw_table_created_at": batch.raw_table_created_at,
        "unique_table_name": batch.unique_table_name,
        "unique_table_created_at": batch.unique_table_created_at,
        "unique_row_count": batch.unique_row_count,
        "imported_at": batch.imported_at,
        "completed_at": batch.completed_at,
        "error_message": batch.error_message,
        "result_summary": batch.result_summary,
        "preview": list(batch.rows.order_by("row_number").values("row_number", "data")[:preview_limit]),
    }
    if include_mappings:
        payload["mappings"] = list(
            batch.mappings.values(
                "id", "mapping_order", "target_table", "target_column", "target_type",
                "suggested_source_column", "source_column", "transform_rule", "casing_type",
                "include", "required", "default_value", "key_role", "notes",
            )
        )
    return payload



def serialize_mapping_template(template):
    return {
        "id": template.id,
        "name": template.name,
        "source_headers": template.source_headers,
        "column_count": template.columns.count(),
        "notes": template.notes,
        "created_at": template.created_at,
        "updated_at": template.updated_at,
    }


@transaction.atomic
def save_mapping_template_from_batch(batch, name):
    clean_name = (name or "").strip()
    if not clean_name:
        raise ValueError("Enter a template name.")

    template, _created = ImportMappingTemplate.objects.update_or_create(
        name=clean_name,
        defaults={"source_headers": batch.headers, "notes": f"Saved from batch {batch.id}: {batch.file_name}"},
    )
    template.columns.all().delete()
    ImportMappingTemplateColumn.objects.bulk_create([
        ImportMappingTemplateColumn(
            template=template,
            mapping_order=mapping.mapping_order,
            target_table=mapping.target_table,
            target_column=mapping.target_column,
            target_type=mapping.target_type,
            source_column=mapping.source_column,
            transform_rule=mapping.transform_rule,
            casing_type=mapping.casing_type,
            include=mapping.include,
            required=mapping.required,
            default_value=mapping.default_value,
            key_role=mapping.key_role,
            notes=mapping.notes,
        )
        for mapping in batch.mappings.order_by("mapping_order")
    ], batch_size=500)
    return template


@transaction.atomic
def apply_mapping_template_to_batch(batch, template):
    valid_headers = set(batch.headers)
    template_columns = list(template.columns.order_by("mapping_order"))
    template_by_target = {
        (column.target_table, column.target_column, column.casing_type): column
        for column in template_columns
    }
    changed = []
    applied_count = 0
    missing_source_columns = []

    for mapping in batch.mappings.order_by("mapping_order"):
        template_column = template_by_target.get((mapping.target_table, mapping.target_column, mapping.casing_type))
        if not template_column:
            continue

        source_column = template_column.source_column if template_column.source_column in valid_headers else ""
        if template_column.source_column and not source_column:
            missing_source_columns.append(template_column.source_column)

        mapping.source_column = source_column
        mapping.include = template_column.include
        mapping.default_value = template_column.default_value
        changed.append(mapping)
        applied_count += 1

    if changed:
        ImportColumnMapping.objects.bulk_update(changed, ["source_column", "include", "default_value"], batch_size=500)
    batch.status = RawImportBatch.STATUS_MAPPED
    batch.error_message = ""
    batch.save(update_fields=["status", "error_message"])

    return {
        "template": serialize_mapping_template(template),
        "applied_count": applied_count,
        "missing_source_columns": sorted(set(missing_source_columns)),
    }

def clean_value(value, mapping):
    if value in (None, ""):
        value = mapping.default_value if mapping.default_value != "" else None
    if value is None:
        return None

    target_type = mapping.target_type.lower()
    if target_type == "text":
        return str(value).strip()
    if target_type in {"double precision", "integer", "bigint"}:
        if isinstance(value, str):
            value = value.strip().replace(",", "")
        if value == "":
            return None
        number = float(value)
        return int(number) if target_type in {"integer", "bigint"} else number
    if target_type == "boolean":
        if isinstance(value, bool):
            return value
        normalized = str(value).strip().casefold()
        if normalized in {"true", "t", "yes", "y", "1"}:
            return True
        if normalized in {"false", "f", "no", "n", "0"}:
            return False
        return None
    if target_type == "date":
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%Y-%m", "%Y/%m"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None
    return value



def parse_production_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_decimal(value):
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def parse_production_decimal(value, required=True):
    if value in (None, ""):
        return (None, False) if required else (Decimal("0"), True)
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)), True
    text = str(value).strip().replace(",", "")
    if not text:
        return (None, False) if required else (Decimal("0"), True)
    try:
        return Decimal(text), True
    except InvalidOperation:
        return None, False


def read_tabular_upload(uploaded_file, requested_sheet=""):
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        if requested_sheet:
            if requested_sheet not in workbook.sheetnames:
                workbook.close()
                raise ValueError(f"Worksheet '{requested_sheet}' was not found.")
            worksheet = workbook[requested_sheet]
        else:
            worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        sheet_name = worksheet.title
        workbook.close()
        if not rows:
            raise ValueError("The selected worksheet is empty.")
        headers = [str(value or "").strip() for value in rows[0]]
        return headers, rows[1:], sheet_name
    if suffix == ".csv":
        uploaded_file.seek(0)
        text = uploaded_file.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            raise ValueError("The CSV file is empty.")
        headers = [str(value or "").strip() for value in rows[0]]
        return headers, rows[1:], "CSV"
    raise ValueError("Upload a .csv, .xlsx, or .xlsm production file.")


PRODUCTION_REQUIRED_FIELDS = ("base_uwi", "production_date", "daily_oil", "daily_water", "daily_gas")
PRODUCTION_MAPPING_FIELDS = (
    {"target": "base_uwi", "label": "base_uwi", "type": "text", "required": True, "rule": "Well ID common key"},
    {"target": "production_date", "label": "date", "type": "date", "required": True, "rule": "Daily production date"},
    {"target": "daily_oil", "label": "oil", "type": "double precision", "required": True, "rule": "Daily oil volume"},
    {"target": "daily_water", "label": "water", "type": "double precision", "required": True, "rule": "Daily water volume"},
    {"target": "daily_gas", "label": "gas", "type": "double precision", "required": True, "rule": "Daily gas volume"},
    {"target": "fluid", "label": "fluid", "type": "double precision", "required": False, "rule": "Imported when present; otherwise oil + water"},
    {"target": "GOR", "label": "GOR", "type": "double precision", "required": False, "rule": "Imported when present; otherwise gas / oil"},
    {"target": "GLR", "label": "GLR", "type": "double precision", "required": False, "rule": "Imported when present; otherwise gas / fluid"},
    {"target": "oilCut", "label": "oilCut", "type": "double precision", "required": False, "rule": "Imported when present; otherwise oil / fluid"},
    {"target": "tbg_psi", "label": "tbg_psi", "type": "double precision", "required": False, "rule": "Tubing pressure (psi)"},
    {"target": "csg_psi", "label": "csg_psi", "type": "double precision", "required": False, "rule": "Casing pressure (psi)"},
    {"target": "chokeSize", "label": "choke size", "type": "double precision", "required": False, "rule": "choke size"},
    {"target": "pumpDEPTH", "label": "pumpDEPTH", "type": "double precision", "required": False, "rule": "pumpDEPTH"},
    {"target": "pumpDia", "label": "pumpDia", "type": "double precision", "required": False, "rule": "pumpDia"},   
)


def suggest_production_columns(headers):
    normalized = {header_key(header): index for index, header in enumerate(headers)}
    mapping = {}
    for target, aliases in PRODUCTION_COLUMN_ALIASES.items():
        for alias in aliases:
            key = header_key(alias)
            if key in normalized:
                mapping[target] = headers[normalized[key]]
                break
    return mapping


def resolve_production_column_map(headers, source_mapping=None, require_all=True):
    header_indexes = {header: index for index, header in enumerate(headers)}
    normalized = {header_key(header): header for header in headers}
    suggested = suggest_production_columns(headers)
    resolved = {}
    source_mapping = source_mapping or {}
    for field in PRODUCTION_MAPPING_FIELDS:
        target = field["target"]
        configured = str(source_mapping.get(target) or "").strip()
        source_header = ""
        if configured in header_indexes:
            source_header = configured
        elif configured and header_key(configured) in normalized:
            source_header = normalized[header_key(configured)]
        elif target in suggested:
            source_header = suggested[target]
        if source_header:
            resolved[target] = header_indexes[source_header]

    if require_all:
        missing = [column for column in PRODUCTION_REQUIRED_FIELDS if column not in resolved]
        if missing:
            raise ValueError("Production import is missing required mappings: " + ", ".join(missing))
    return resolved


def detect_production_columns(headers):
    return resolve_production_column_map(headers)


def ensure_production_tables(cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS production_daily (
            base_uwi text NOT NULL,
            production_date date NOT NULL,
            daily_oil double precision NOT NULL DEFAULT 0,
            daily_water double precision NOT NULL DEFAULT 0,
            daily_gas double precision NOT NULL DEFAULT 0,
            fluid double precision NOT NULL DEFAULT 0,
            GOR double precision NOT NULL DEFAULT 0,
            GLR double precision NOT NULL DEFAULT 0,
            oilCut double precision NOT NULL DEFAULT 0,
            tbg_psi double precision NOT NULL DEFAULT 0,
            csg_psi double precision NOT NULL DEFAULT 0,
            chokeSize double precision NOT NULL DEFAULT 0,
            pumpDEPTH double precision NOT NULL DEFAULT 0,
            pumpDia double precision NOT NULL DEFAULT 0,
            source_file text,
            imported_at timestamptz NOT NULL DEFAULT now()
        )
        """
    )
    cursor.execute(
    """
    ALTER TABLE production_daily
    ADD COLUMN IF NOT EXISTS tbg_psi double precision NOT NULL DEFAULT 0,
    ADD COLUMN IF NOT EXISTS csg_psi double precision NOT NULL DEFAULT 0,
    ADD COLUMN IF NOT EXISTS GOR double precision NOT NULL DEFAULT 0,
    ADD COLUMN IF NOT EXISTS GLR double precision NOT NULL DEFAULT 0
    """
    )
    cursor.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS production_daily_base_uwi_date_idx
        ON production_daily (base_uwi, production_date)
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS production_monthly (
            base_uwi text NOT NULL,
            production_month date NOT NULL,
            monthly_oil double precision NOT NULL DEFAULT 0,
            monthly_water double precision NOT NULL DEFAULT 0,
            monthly_gas double precision NOT NULL DEFAULT 0,
            monthly_fluid double precision NOT NULL DEFAULT 0,
            monthly_averageGOR double precision NOT NULL DEFAULT 0,
            monthly_averageGWR double precision NOT NULL DEFAULT 0,
            cumulative_oil double precision NOT NULL DEFAULT 0,
            cumulative_water double precision NOT NULL DEFAULT 0,
            cumulative_gas double precision NOT NULL DEFAULT 0,
            cumulative_fluid double precision NOT NULL DEFAULT 0,
            updated_at timestamptz NOT NULL DEFAULT now()
        )
        """
    )
    cursor.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS production_monthly_base_uwi_month_idx
        ON production_monthly (base_uwi, production_month)
        """
    )


def serialize_production_row(row):
    return {
        "base_uwi": row["base_uwi"],
        "production_date": row["production_date"].isoformat(),
        "daily_oil": float(row["daily_oil"]),
        "daily_water": float(row["daily_water"]),
        "daily_gas": float(row["daily_gas"]),
        "fluid": float(row["fluid"]),
        "tbg_psi": float(row["tbg_psi"]),
        "csg_psi": float(row["csg_psi"]),
        "GOR": float(row["GOR"]),
        "GLR": float(row["GLR"]),
    }


def production_source_value(values, column_map, target):
    index = column_map.get(target)
    if index is None or index >= len(values):
        return None
    return values[index]


def build_production_rows(headers, source_rows, source_mapping=None, require_all=True, limit=None):
    column_map = resolve_production_column_map(headers, source_mapping, require_all=require_all)
    parsed_rows = []
    skipped_rows = 0

    for values in source_rows:
        values = list(values)
        if not any(value not in (None, "") for value in values):
            continue
        base_uwi = str(production_source_value(values, column_map, "base_uwi") or "").strip()
        production_date = parse_production_date(production_source_value(values, column_map, "production_date"))
        if not base_uwi or not production_date:
            skipped_rows += 1
            continue
        daily_oil, oil_valid = parse_production_decimal(production_source_value(values, column_map, "daily_oil"))
        daily_water, water_valid = parse_production_decimal(production_source_value(values, column_map, "daily_water"))
        daily_gas, gas_valid = parse_production_decimal(production_source_value(values, column_map, "daily_gas"))
        fluid_source = production_source_value(values, column_map, "fluid")
        if "fluid" in column_map and fluid_source not in (None, ""):
            fluid, fluid_valid = parse_production_decimal(fluid_source)
        else:
            fluid_valid = True
            fluid = daily_oil + daily_water
        if not (oil_valid and water_valid and gas_valid and fluid_valid):
            skipped_rows += 1
            continue

        optional = {}
        optional_valid = True
        for target in ("tbg_psi", "csg_psi"):
            value, valid = parse_production_decimal(
                production_source_value(values, column_map, target), required=False
            )
            optional[target] = value
            optional_valid = optional_valid and valid

        ratio_defaults = {
            "GOR": daily_gas / daily_oil if daily_oil else Decimal("0"),
            "GLR": daily_gas / fluid if fluid else Decimal("0"),
        }
        for target, calculated_value in ratio_defaults.items():
            source_value = production_source_value(values, column_map, target)
            if target in column_map and source_value not in (None, ""):
                value, valid = parse_production_decimal(source_value)
                optional[target] = value
                optional_valid = optional_valid and valid
            else:
                optional[target] = calculated_value

        if not optional_valid:
            skipped_rows += 1
            continue
        parsed_rows.append({
            "base_uwi": base_uwi, "production_date": production_date,
            "daily_oil": daily_oil, "daily_water": daily_water,
            "daily_gas": daily_gas, "fluid": fluid, **optional,
        })
        if limit and len(parsed_rows) >= limit:
            break
    return parsed_rows, skipped_rows, column_map


def inspect_production_file(uploaded_file, requested_sheet=""):
    headers, source_rows, sheet_name = read_tabular_upload(uploaded_file, requested_sheet)
    suggested = suggest_production_columns(headers)
    return {
        "file_name": uploaded_file.name,
        "sheet_name": sheet_name,
        "headers": headers,
        "row_count": sum(1 for row in source_rows if any(value not in (None, "") for value in row)),
        "fields": list(PRODUCTION_MAPPING_FIELDS),
        "suggested_mappings": suggested,
        "preview": [
            {headers[index] or f"column_{index + 1}": json_value(value) for index, value in enumerate(row[:len(headers)])}
            for row in source_rows[:20]
        ],
    }


def classify_production_rows(cursor, parsed_rows, mapped_fields):
    ensure_production_tables(cursor)
    wells = sorted({row["base_uwi"] for row in parsed_rows})
    dates = sorted({row["production_date"] for row in parsed_rows})
    existing = {}
    if wells and dates:
        cursor.execute(
            """SELECT base_uwi, production_date, daily_oil, daily_water, daily_gas, fluid,
                      tbg_psi, csg_psi, GOR, GLR
               FROM production_daily WHERE base_uwi = ANY(%s) AND production_date = ANY(%s)""",
            [wells, dates],
        )
        columns = ("base_uwi", "production_date", "daily_oil", "daily_water", "daily_gas", "fluid",
                   "tbg_psi", "csg_psi", "GOR", "GLR")
        existing = {(record[0], record[1]): dict(zip(columns, record)) for record in cursor.fetchall()}
    new_rows, duplicate_rows, conflicts = [], [], []
    compared_fields = [field for field in mapped_fields if field not in {"base_uwi", "production_date"}]
    for uploaded in parsed_rows:
        stored = existing.get((uploaded["base_uwi"], uploaded["production_date"]))
        if stored is None:
            new_rows.append(uploaded)
            continue
        differences = [field for field in compared_fields if Decimal(str(stored[field])) != uploaded[field]]
        if not differences:
            duplicate_rows.append(uploaded)
        else:
            conflicts.append({
                "different_fields": differences,
                "existing": serialize_production_row(stored),
                "uploaded": serialize_production_row(uploaded),
            })
    return new_rows, duplicate_rows, conflicts


def preview_production_file(uploaded_file, source_mapping, requested_sheet=""):
    headers, source_rows, sheet_name = read_tabular_upload(uploaded_file, requested_sheet)
    parsed_rows, invalid_count, column_map = build_production_rows(headers, source_rows, source_mapping)
    with connection.cursor() as cursor:
        new_rows, duplicates, conflicts = classify_production_rows(cursor, parsed_rows, column_map)
    return {
        "file_name": uploaded_file.name, "sheet_name": sheet_name,
        "daily_table_name": PRODUCTION_DAILY_TABLE, "monthly_table_name": PRODUCTION_MONTHLY_TABLE,
        "new_row_count": len(new_rows), "duplicate_row_count": len(duplicates),
        "conflict_row_count": len(conflicts), "invalid_row_count": invalid_count,
        "only_duplicates": bool(duplicates) and not new_rows and not conflicts and not invalid_count,
        "mapped_columns": {target: headers[index] for target, index in column_map.items()},
        "preview": [serialize_production_row(row) for row in new_rows[:20]],
        "conflicts": conflicts[:100],
    }


@transaction.atomic
def import_production_file(uploaded_file, requested_sheet="", replace_conflicts=False, source_mapping=None):
    headers, source_rows, sheet_name = read_tabular_upload(uploaded_file, requested_sheet)
    parsed_rows, invalid_count, column_map = build_production_rows(headers, source_rows, source_mapping)
    imported_at = timezone.now()
    with connection.cursor() as cursor:
        new_rows, duplicates, conflicts = classify_production_rows(cursor, parsed_rows, column_map)
        if duplicates and not new_rows and not conflicts and not invalid_count:
            raise ValueError("Every valid uploaded record is an exact duplicate; there is no new data to add.")
        replacement_rows = []
        if replace_conflicts:
            for item in conflicts:
                row = item["uploaded"]
                replacement_rows.append({
                    "base_uwi": row["base_uwi"], "production_date": parse_production_date(row["production_date"]),
                    "daily_oil": Decimal(str(row["daily_oil"])), "daily_water": Decimal(str(row["daily_water"])),
                    "daily_gas": Decimal(str(row["daily_gas"])), "fluid": Decimal(str(row["fluid"])),
                    "tbg_psi": Decimal(str(row["tbg_psi"])), "csg_psi": Decimal(str(row["csg_psi"])),
                    "GOR": Decimal(str(row["GOR"])), "GLR": Decimal(str(row["GLR"])),
                })
        approved_rows = new_rows + replacement_rows
        imported_base_uwis = sorted({row["base_uwi"] for row in approved_rows})
        insert_daily = """
            INSERT INTO production_daily
                (base_uwi, production_date, daily_oil, daily_water, daily_gas, fluid, tbg_psi, csg_psi, GOR, GLR, source_file, imported_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (base_uwi, production_date) DO UPDATE SET
                daily_oil=EXCLUDED.daily_oil, daily_water=EXCLUDED.daily_water,
                daily_gas=EXCLUDED.daily_gas, fluid=EXCLUDED.fluid,
                tbg_psi=EXCLUDED.tbg_psi, csg_psi=EXCLUDED.csg_psi, GOR=EXCLUDED.GOR, GLR=EXCLUDED.GLR,
                source_file=EXCLUDED.source_file, imported_at=EXCLUDED.imported_at
        """
        cursor.executemany(insert_daily, [[
            row["base_uwi"], row["production_date"], row["daily_oil"], row["daily_water"],
            row["daily_gas"], row["fluid"], row["tbg_psi"], row["csg_psi"], row["GOR"], row["GLR"],
            uploaded_file.name, imported_at
        ] for row in approved_rows])
        monthly_row_count = 0
        if imported_base_uwis:
            cursor.execute("DELETE FROM production_monthly WHERE base_uwi = ANY(%s)", [imported_base_uwis])
            cursor.execute(
                """
                INSERT INTO production_monthly (
                    base_uwi, production_month, monthly_oil, monthly_water, monthly_gas, monthly_fluid,
                    cumulative_oil, cumulative_water, cumulative_gas, cumulative_fluid, updated_at
                )
                WITH monthly AS (
                    SELECT base_uwi, date_trunc('month', production_date)::date production_month,
                           SUM(daily_oil) monthly_oil, SUM(daily_water) monthly_water,
                           SUM(daily_gas) monthly_gas, SUM(fluid) monthly_fluid
                    FROM production_daily WHERE base_uwi = ANY(%s)
                    GROUP BY base_uwi, date_trunc('month', production_date)::date
                )
                SELECT base_uwi, production_month, monthly_oil, monthly_water, monthly_gas, monthly_fluid,
                       SUM(monthly_oil) OVER (PARTITION BY base_uwi ORDER BY production_month),
                       SUM(monthly_water) OVER (PARTITION BY base_uwi ORDER BY production_month),
                       SUM(monthly_gas) OVER (PARTITION BY base_uwi ORDER BY production_month),
                       SUM(monthly_fluid) OVER (PARTITION BY base_uwi ORDER BY production_month), now()
                FROM monthly ORDER BY base_uwi, production_month
                """, [imported_base_uwis])
            cursor.execute("SELECT COUNT(*) FROM production_monthly WHERE base_uwi = ANY(%s)", [imported_base_uwis])
            monthly_row_count = cursor.fetchone()[0]
    return {
        "file_name": uploaded_file.name, "sheet_name": sheet_name,
        "daily_table_name": PRODUCTION_DAILY_TABLE, "monthly_table_name": PRODUCTION_MONTHLY_TABLE,
        "daily_row_count": len(approved_rows), "monthly_row_count": monthly_row_count,
        "inserted_row_count": len(new_rows), "replaced_row_count": len(replacement_rows),
        "skipped_row_count": len(duplicates) + (0 if replace_conflicts else len(conflicts)),
        "rejected_row_count": invalid_count, "well_count": len(imported_base_uwis),
        "replace_conflicts": replace_conflicts,
        "mapped_columns": {target: headers[index] for target, index in column_map.items()},
        "preview": [serialize_production_row(row) for row in approved_rows[:20]],
    }
INJECTION_REQUIRED_FIELDS = ("base_uwi", "injection_date")
INJECTION_MAPPING_FIELDS = (
    {"target": "base_uwi", "label": "base_uwi", "type": "text", "required": True, "rule": "Well ID common key"},
    {"target": "injection_date", "label": "date", "type": "date", "required": True, "rule": "Daily injection date"},
    {"target": "daily_water", "label": "water", "type": "double precision", "required": False, "rule": "Injected water volume (bbl)"},
    {"target": "daily_gas", "label": "gas", "type": "double precision", "required": False, "rule": "Injected gas volume (mcf)"},
    {"target": "daily_steam", "label": "steam", "type": "double precision", "required": False, "rule": "Injected steam volume (bbl)"},
    {"target": "injection_pressure", "label": "pressure", "type": "double precision", "required": False, "rule": "Injection pressure"},
)
INJECTION_COLUMN_ALIASES = {
    "base_uwi": PRODUCTION_COLUMN_ALIASES["base_uwi"],
    "injection_date": ("date", "injection_date", "injection date", "inject_date", "inject date", "day"),
    "daily_water": ("daily_water", "daily water", "water", "water_injected", "water injected", "water_bbl", "water bbl"),
    "daily_gas": ("daily_gas", "daily gas", "gas", "gas_injected", "gas injected", "gas_mcf", "gas mcf"),
    "daily_steam": ("daily_steam", "daily steam", "steam", "steam_injected", "steam injected", "steam_bbl", "steam bbl"),
    "injection_pressure": ("injection_pressure", "injection pressure", "pressure", "pressure_psi", "pressure psi"),
}


def suggest_injection_columns(headers):
    normalized = {header_key(header): index for index, header in enumerate(headers)}
    mapping = {}
    for target, aliases in INJECTION_COLUMN_ALIASES.items():
        for alias in aliases:
            key = header_key(alias)
            if key in normalized:
                mapping[target] = headers[normalized[key]]
                break
    return mapping


def resolve_injection_column_map(headers, source_mapping=None):
    header_indexes = {header: index for index, header in enumerate(headers)}
    normalized = {header_key(header): header for header in headers}
    suggested = suggest_injection_columns(headers)
    resolved = {}
    for field in INJECTION_MAPPING_FIELDS:
        target = field["target"]
        configured = str((source_mapping or {}).get(target) or "").strip()
        source_header = configured if configured in header_indexes else normalized.get(header_key(configured), "")
        if not source_header:
            source_header = suggested.get(target, "")
        if source_header:
            resolved[target] = header_indexes[source_header]
    missing = [field for field in INJECTION_REQUIRED_FIELDS if field not in resolved]
    if missing:
        raise ValueError("Injection import is missing required mappings: " + ", ".join(missing))
    if not any(field in resolved for field in ("daily_water", "daily_gas", "daily_steam")):
        raise ValueError("Map at least one injected-volume field: water, gas, or steam.")
    return resolved


def injection_source_value(values, column_map, target):
    index = column_map.get(target)
    return values[index] if index is not None and index < len(values) else None


def parse_injection_number(value, label):
    if value in (None, ""):
        return Decimal("0"), None
    try:
        number = Decimal(str(value).strip().replace(",", ""))
    except InvalidOperation:
        return Decimal("0"), f"{label} is not numeric"
    if number < 0:
        return number, f"{label} cannot be negative"
    return number, None


def build_injection_rows(headers, source_rows, source_mapping=None, limit=None):
    column_map = resolve_injection_column_map(headers, source_mapping)
    parsed_rows, validation_errors = [], []
    for row_number, source_values in enumerate(source_rows, start=2):
        values = list(source_values)
        if not any(value not in (None, "") for value in values):
            continue
        base_uwi = str(injection_source_value(values, column_map, "base_uwi") or "").strip()
        injection_date = parse_production_date(injection_source_value(values, column_map, "injection_date"))
        errors = []
        if not base_uwi:
            errors.append("base_uwi is required")
        if not injection_date:
            errors.append("date is invalid or missing")
        volumes = {}
        for target, label in (("daily_water", "water"), ("daily_gas", "gas"), ("daily_steam", "steam"), ("injection_pressure", "pressure")):
            volumes[target], error = parse_injection_number(injection_source_value(values, column_map, target), label)
            if error:
                errors.append(error)
        if errors:
            validation_errors.append({"row_number": row_number, "base_uwi": base_uwi, "errors": errors})
            continue
        parsed_rows.append({"base_uwi": base_uwi, "injection_date": injection_date, **volumes})
        if limit and len(parsed_rows) >= limit:
            break
    return parsed_rows, validation_errors, column_map


def serialize_injection_row(row):
    return {
        "base_uwi": row["base_uwi"], "injection_date": row["injection_date"].isoformat(),
        "daily_water": float(row["daily_water"]), "daily_gas": float(row["daily_gas"]),
        "daily_steam": float(row["daily_steam"]), "injection_pressure": float(row["injection_pressure"]),
    }


def classify_injection_rows(cursor, parsed_rows, mapped_fields):
    ensure_injection_tables(cursor)
    wells = sorted({row["base_uwi"] for row in parsed_rows})
    dates = sorted({row["injection_date"] for row in parsed_rows})
    existing = {}
    if wells and dates:
        cursor.execute(
            """SELECT base_uwi, injection_date, daily_water, daily_gas, daily_steam, injection_pressure
               FROM injection_daily WHERE base_uwi = ANY(%s) AND injection_date = ANY(%s)""",
            [wells, dates],
        )
        columns = ("base_uwi", "injection_date", "daily_water", "daily_gas", "daily_steam", "injection_pressure")
        existing = {(record[0], record[1]): dict(zip(columns, record)) for record in cursor.fetchall()}
    new_rows, duplicate_rows, conflicts = [], [], []
    compared_fields = [field for field in mapped_fields if field not in {"base_uwi", "injection_date"}]
    for uploaded in parsed_rows:
        stored = existing.get((uploaded["base_uwi"], uploaded["injection_date"]))
        if stored is None:
            new_rows.append(uploaded)
            continue
        differences = [field for field in compared_fields if Decimal(str(stored[field])) != uploaded[field]]
        if not differences:
            duplicate_rows.append(uploaded)
        else:
            conflicts.append({
                "different_fields": differences,
                "existing": serialize_injection_row(stored),
                "uploaded": serialize_injection_row(uploaded),
            })
    return new_rows, duplicate_rows, conflicts


def inspect_injection_file(uploaded_file, requested_sheet=""):
    headers, source_rows, sheet_name = read_tabular_upload(uploaded_file, requested_sheet)
    sheet_names = [sheet_name]
    if Path(uploaded_file.name).suffix.lower() in {".xlsx", ".xlsm"}:
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
    return {
        "file_name": uploaded_file.name, "sheet_name": sheet_name, "sheet_names": sheet_names,
        "headers": headers, "row_count": sum(1 for row in source_rows if any(value not in (None, "") for value in row)),
        "fields": list(INJECTION_MAPPING_FIELDS), "suggested_mappings": suggest_injection_columns(headers),
        "preview": [{headers[i] or f"column_{i + 1}": json_value(value) for i, value in enumerate(row[:len(headers)])} for row in source_rows[:20]],
    }


def preview_injection_file(uploaded_file, source_mapping, requested_sheet=""):
    headers, source_rows, sheet_name = read_tabular_upload(uploaded_file, requested_sheet)
    rows, errors, column_map = build_injection_rows(headers, source_rows, source_mapping)
    if not rows:
        raise ValueError("No valid injection rows were found. Review the validation errors and mappings.")
    return {
        "file_name": uploaded_file.name, "sheet_name": sheet_name,
        "daily_table_name": INJECTION_DAILY_TABLE, "monthly_table_name": INJECTION_MONTHLY_TABLE,
        "valid_row_count": len(rows), "invalid_row_count": len(errors), "validation_errors": errors[:50],
        "mapped_columns": {target: headers[index] for target, index in column_map.items()},
        "preview": [serialize_injection_row(row) for row in rows[:20]],
    }


def ensure_injection_tables(cursor):
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS injection_daily (
            base_uwi text NOT NULL, injection_date date NOT NULL,
            daily_water double precision NOT NULL DEFAULT 0, daily_gas double precision NOT NULL DEFAULT 0,
            daily_steam double precision NOT NULL DEFAULT 0, injection_pressure double precision NOT NULL DEFAULT 0,
            source_file text, imported_at timestamptz NOT NULL DEFAULT now(),
            UNIQUE (base_uwi, injection_date)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS injection_monthly (
            base_uwi text NOT NULL, injection_month date NOT NULL,
            monthly_water double precision NOT NULL DEFAULT 0, monthly_gas double precision NOT NULL DEFAULT 0,
            monthly_steam double precision NOT NULL DEFAULT 0, cumulative_water double precision NOT NULL DEFAULT 0,
            cumulative_gas double precision NOT NULL DEFAULT 0, cumulative_steam double precision NOT NULL DEFAULT 0,
            updated_at timestamptz NOT NULL DEFAULT now(), UNIQUE (base_uwi, injection_month)
        )
    """)


@transaction.atomic
def import_injection_file(uploaded_file, requested_sheet="", replace_conflicts=False, source_mapping=None):
    headers, source_rows, sheet_name = read_tabular_upload(uploaded_file, requested_sheet)
    parsed_rows, errors, column_map = build_injection_rows(headers, source_rows, source_mapping)
    if not parsed_rows:
        raise ValueError("No valid injection rows were found. Review the file and mappings.")
    imported_at = timezone.now()
    with connection.cursor() as cursor:
        new_rows, duplicates, conflicts = classify_injection_rows(cursor, parsed_rows, column_map)
        if duplicates and not new_rows and not conflicts and not errors:
            raise ValueError("Every valid uploaded record is an exact duplicate; there is no new data to add.")
        replacement_rows = []
        if replace_conflicts:
            for item in conflicts:
                row = item["uploaded"]
                replacement_rows.append({
                    "base_uwi": row["base_uwi"], "injection_date": parse_production_date(row["injection_date"]),
                    "daily_water": Decimal(str(row["daily_water"])), "daily_gas": Decimal(str(row["daily_gas"])),
                    "daily_steam": Decimal(str(row["daily_steam"])), "injection_pressure": Decimal(str(row["injection_pressure"])),
                })
        approved_rows = new_rows + replacement_rows
        base_uwis = sorted({row["base_uwi"] for row in approved_rows})
        cursor.executemany("""
            INSERT INTO injection_daily
                (base_uwi, injection_date, daily_water, daily_gas, daily_steam, injection_pressure, source_file, imported_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (base_uwi, injection_date) DO UPDATE SET
                daily_water=EXCLUDED.daily_water, daily_gas=EXCLUDED.daily_gas,
                daily_steam=EXCLUDED.daily_steam, injection_pressure=EXCLUDED.injection_pressure,
                source_file=EXCLUDED.source_file, imported_at=EXCLUDED.imported_at
        """, [[row["base_uwi"], row["injection_date"], row["daily_water"], row["daily_gas"], row["daily_steam"], row["injection_pressure"], uploaded_file.name, imported_at] for row in approved_rows])
        monthly_count = 0
        if base_uwis:
            cursor.execute("DELETE FROM injection_monthly WHERE base_uwi = ANY(%s)", [base_uwis])
            cursor.execute("""
                INSERT INTO injection_monthly (
                    base_uwi, injection_month, monthly_water, monthly_gas, monthly_steam,
                    cumulative_water, cumulative_gas, cumulative_steam, updated_at
                )
                WITH monthly AS (
                    SELECT base_uwi, date_trunc('month', injection_date)::date injection_month,
                           SUM(daily_water) monthly_water, SUM(daily_gas) monthly_gas, SUM(daily_steam) monthly_steam
                    FROM injection_daily WHERE base_uwi = ANY(%s)
                    GROUP BY base_uwi, date_trunc('month', injection_date)::date
                )
                SELECT base_uwi, injection_month, monthly_water, monthly_gas, monthly_steam,
                       SUM(monthly_water) OVER (PARTITION BY base_uwi ORDER BY injection_month),
                       SUM(monthly_gas) OVER (PARTITION BY base_uwi ORDER BY injection_month),
                       SUM(monthly_steam) OVER (PARTITION BY base_uwi ORDER BY injection_month), now()
                FROM monthly
            """, [base_uwis])
            cursor.execute("SELECT COUNT(*) FROM injection_monthly WHERE base_uwi = ANY(%s)", [base_uwis])
            monthly_count = cursor.fetchone()[0]
    return {
        "file_name": uploaded_file.name, "sheet_name": sheet_name,
        "daily_table_name": INJECTION_DAILY_TABLE, "monthly_table_name": INJECTION_MONTHLY_TABLE,
        "daily_row_count": len(approved_rows), "monthly_row_count": monthly_count,
        "inserted_row_count": len(new_rows), "replaced_row_count": len(replacement_rows),
        "skipped_row_count": len(duplicates) + (0 if replace_conflicts else len(conflicts)),
        "valid_row_count": len(parsed_rows), "invalid_row_count": len(errors), "validation_errors": errors[:50],
        "well_count": len(base_uwis), "replace_conflicts": replace_conflicts,
        "mapped_columns": {target: headers[index] for target, index in column_map.items()},
        "preview": [serialize_injection_row(row) for row in approved_rows[:20]],
    }

def mapping_value(raw_row, mapping, fallback_raw_id=None, source_file="", imported_at=None):
    if mapping.source_column:
        value = raw_row.get(mapping.source_column)
    elif mapping.target_column == "raw_id":
        value = fallback_raw_id
    elif mapping.target_column == "source_file":
        value = source_file
    elif mapping.target_column == "import_timestamp":
        value = imported_at.isoformat() if imported_at else None
    else:
        value = None
    try:
        return clean_value(value, mapping)
    except (TypeError, ValueError, OverflowError):
        return None


def validate_identifier(value):
    if not IDENTIFIER_RE.fullmatch(value or ""):
        raise ValueError(f"Unsafe database identifier: {value!r}")
    return value


def ensure_target_table(cursor, table_name, column_types):
    validate_identifier(table_name)
    cursor.execute(sql.SQL("CREATE TABLE IF NOT EXISTS {} ()").format(sql.Identifier(table_name)))
    for column_name, target_type in column_types.items():
        validate_identifier(column_name)
        sql_type = TYPE_MAP.get((target_type or "text").lower(), "text")
        cursor.execute(
            sql.SQL("ALTER TABLE {} ADD COLUMN IF NOT EXISTS {} {}").format(
                sql.Identifier(table_name), sql.Identifier(column_name), sql.SQL(sql_type)
            )
        )


def merge_mapping_value(target, mapping, value):
    if not mapping.target_column:
        return
    if target.get(mapping.target_column) in (None, "") and value not in (None, ""):
        target[mapping.target_column] = value
    elif mapping.target_column not in target:
        target[mapping.target_column] = value


def insert_rows(cursor, table_name, rows):
    if not rows:
        return 0
    columns = list(rows[0].keys())
    statement = sql.SQL("INSERT INTO {} ({}) VALUES ({})").format(
        sql.Identifier(table_name),
        sql.SQL(", ").join(map(sql.Identifier, columns)),
        sql.SQL(", ").join(sql.Placeholder() for _ in columns),
    )
    cursor.executemany(statement, [[row.get(column) for column in columns] for row in rows])
    return len(rows)


def upsert_rows_by_base_uwi(cursor, table_name, rows):
    """Insert rows, first clearing any prior rows for the same wells.

    Re-importing a well replaces its rows in this table without disturbing
    rows belonging to other wells, so append imports don't need a full
    TRUNCATE to avoid duplicate accumulation on repeat uploads.
    """
    if not rows:
        return 0
    if "base_uwi" in rows[0]:
        base_uwis = sorted({row["base_uwi"] for row in rows if row.get("base_uwi")})
        if base_uwis:
            cursor.execute(
                sql.SQL("DELETE FROM {} WHERE {} = ANY(%s)").format(
                    sql.Identifier(table_name), sql.Identifier("base_uwi")
                ),
                [base_uwis],
            )
    return insert_rows(cursor, table_name, rows)


def build_output_groups(mappings):
    common = [mapping for mapping in mappings if mapping.target_table == "__common_keys__" and mapping.include and mapping.target_column]
    tables = {}
    for mapping in mappings:
        if not mapping.include or mapping.target_table == "__common_keys__" or not mapping.target_column:
            continue
        tables.setdefault(mapping.target_table, []).append(mapping)
    return common, tables


def preview_mapped_batch(batch, limit=20):
    mappings = list(batch.mappings.order_by("mapping_order"))
    missing_required = [
        mapping.target_column for mapping in mappings
        if mapping.include and mapping.required and not mapping.source_column
        and not mapping.default_value
        and mapping.target_column not in {"raw_id", "source_file", "import_timestamp"}
    ]
    if missing_required:
        raise ValueError("Required mappings are unresolved: " + ", ".join(missing_required))

    common, tables = build_output_groups(mappings)
    previews = {
        table_name: {
            "table_name": table_name,
            "headers": list(dict.fromkeys(
                [mapping.target_column for mapping in common]
                + (["casing_type"] if table_name == "well_casing" else [])
                + [mapping.target_column for mapping in table_mappings]
            )),
            "rows": [],
        }
        for table_name, table_mappings in tables.items()
    }

    for raw in batch.rows.order_by("row_number").iterator(chunk_size=100):
        base_values = {}
        for mapping in common:
            merge_mapping_value(base_values, mapping, mapping_value(raw.data, mapping, raw.id, batch.file_name, batch.imported_at))

        for table_name, table_mappings in tables.items():
            preview = previews[table_name]
            if len(preview["rows"]) >= limit:
                continue
            if table_name == "well_casing":
                casing_groups = {}
                for mapping in table_mappings:
                    casing_groups.setdefault(mapping.casing_type or "unknown", []).append(mapping)
                for casing_index, (casing_type, casing_mappings) in enumerate(casing_groups.items(), start=1):
                    if len(preview["rows"]) >= limit:
                        break
                    output = dict(base_values)
                    has_specific_value = False
                    for mapping in casing_mappings:
                        value = mapping_value(raw.data, mapping, raw.id, batch.file_name, batch.imported_at)
                        merge_mapping_value(output, mapping, value)
                        has_specific_value = has_specific_value or value not in (None, "")
                    if not has_specific_value:
                        continue
                    output["casing_type"] = casing_type
                    if "raw_id" in output:
                        output["raw_id"] = raw.id * 100 + casing_index
                    preview["rows"].append({header: json_value(output.get(header)) for header in preview["headers"]})
            else:
                output = dict(base_values)
                has_specific_value = False
                for mapping in table_mappings:
                    value = mapping_value(raw.data, mapping, raw.id, batch.file_name, batch.imported_at)
                    merge_mapping_value(output, mapping, value)
                    has_specific_value = has_specific_value or value not in (None, "")
                if has_specific_value:
                    preview["rows"].append({header: json_value(output.get(header)) for header in preview["headers"]})

        if previews and all(len(preview["rows"]) >= limit for preview in previews.values()):
            break

    return list(previews.values())

@transaction.atomic
def split_batch(batch, replace_existing=True):
    mappings = list(batch.mappings.order_by("mapping_order"))
    missing_required = [mapping.target_column for mapping in mappings if mapping.include and mapping.required and not mapping.source_column and not mapping.default_value and mapping.target_column not in {"raw_id", "source_file", "import_timestamp"}]
    if missing_required:
        raise ValueError("Required mappings are unresolved: " + ", ".join(missing_required))

    common, tables = build_output_groups(mappings)
    summary = {}
    imported_at = timezone.now()

    with connection.cursor() as cursor:
        for table_name, table_mappings in tables.items():
            if table_name == "well_header":
                for column_name in WELL_HEADER_MOVED_COLUMNS:
                    cursor.execute(sql.SQL("ALTER TABLE {} DROP COLUMN IF EXISTS {} CASCADE").format(
                        sql.Identifier(table_name), sql.Identifier(column_name)
                    ))
            column_types = {mapping.target_column: mapping.target_type for mapping in common + table_mappings}
            if table_name == "well_casing":
                column_types["casing_type"] = "text"
            ensure_target_table(cursor, table_name, column_types)
            if table_name == "well_header":
                cursor.execute("DELETE FROM well_header WHERE NULLIF(BTRIM(base_uwi), '') IS NULL")
                cursor.execute("""
                    DELETE FROM well_header older
                    USING well_header newer
                    WHERE older.base_uwi = newer.base_uwi
                      AND older.ctid < newer.ctid
                """)
                cursor.execute("""
                    CREATE UNIQUE INDEX IF NOT EXISTS well_header_base_uwi_unique_idx
                    ON well_header (base_uwi)
                """)
            if replace_existing:
                cursor.execute(sql.SQL("TRUNCATE TABLE {}").format(sql.Identifier(table_name)))

        buffers = {table_name: [] for table_name in tables}
        unique_well_headers = {}
        for raw in batch.rows.order_by("row_number").iterator(chunk_size=500):
            base_values = {}
            for mapping in common:
                merge_mapping_value(base_values, mapping, mapping_value(raw.data, mapping, raw.id, batch.file_name, imported_at))

            for table_name, table_mappings in tables.items():
                if table_name == "well_casing":
                    casing_groups = {}
                    for mapping in table_mappings:
                        casing_groups.setdefault(mapping.casing_type or "unknown", []).append(mapping)
                    for casing_index, (casing_type, casing_mappings) in enumerate(casing_groups.items(), start=1):
                        output = dict(base_values)
                        has_specific_value = False
                        for mapping in casing_mappings:
                            value = mapping_value(raw.data, mapping, raw.id, batch.file_name, imported_at)
                            merge_mapping_value(output, mapping, value)
                            has_specific_value = has_specific_value or value not in (None, "")
                        if not has_specific_value:
                            continue
                        output["casing_type"] = casing_type
                        if "raw_id" in output:
                            output["raw_id"] = raw.id * 100 + casing_index
                        buffers[table_name].append(output)
                else:
                    output = dict(base_values)
                    has_specific_value = False
                    for mapping in table_mappings:
                        value = mapping_value(raw.data, mapping, raw.id, batch.file_name, imported_at)
                        merge_mapping_value(output, mapping, value)
                        has_specific_value = has_specific_value or value not in (None, "")
                    if has_specific_value or table_name == "well_header":
                        if table_name == "well_header":
                            base_uwi = str(output.get("base_uwi") or "").strip()
                            if base_uwi:
                                output["base_uwi"] = base_uwi
                                unique_well_headers[base_uwi] = output
                        else:
                            buffers[table_name].append(output)

                if table_name != "well_header" and len(buffers[table_name]) >= 500:
                    summary[table_name] = summary.get(table_name, 0) + insert_rows(cursor, table_name, buffers[table_name])
                    buffers[table_name] = []

        for table_name, rows in buffers.items():
            if table_name != "well_header":
                summary[table_name] = summary.get(table_name, 0) + insert_rows(cursor, table_name, rows)
        if "well_header" in tables:
            summary["well_header"] = upsert_rows_by_base_uwi(cursor, "well_header", list(unique_well_headers.values()))

    batch.status = RawImportBatch.STATUS_COMPLETED
    batch.completed_at = timezone.now()
    batch.error_message = ""
    batch.result_summary = summary
    batch.save(update_fields=["status", "completed_at", "error_message", "result_summary"])
    return summary
